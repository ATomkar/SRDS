import openpyxl


class Data:

    @staticmethod
    def getTestData(file_path):
        lst = []
        try:
            # Load the workbook and sheet
            book = openpyxl.load_workbook(file_path)
            sheet = book.active

            # Iterate through each row
            for i in range(2, sheet.max_row + 1):  # Start from row 2 to skip the header row
                username = sheet.cell(row=i, column=2).value  # Assuming the username is in the second column
                if username:  # Ensure the username is not empty
                    # Extract data for the current row
                    data = {}
                    for j in range(1, sheet.max_column + 1):  # Loop through all columns
                        key = sheet.cell(row=1, column=j).value  # Column headers (Testcase_no., Username, etc.)
                        value = sheet.cell(row=i, column=j).value  # Data from the row
                        data[key] = value
                    lst.append(data)

            return lst

        except FileNotFoundError:
            print(f"Error: The file at {file_path} was not found.")
            return []  # Return an empty list in case of failure
        except Exception as e:
            print(f"An error occurred: {e}")
            return []  # Return an empty list in case of any error
